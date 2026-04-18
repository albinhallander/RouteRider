#!/usr/bin/env python3
"""
RouteRider — Merge
Slår ihop routerider_foretag.xlsx (HQ-data från Allabolag)
med routerider_lager.json (lager/fabriker från warehouse_finder.py)
till ett enda Excel-ark med flaggan pickup_type: HQ | Lager | Fabrik | Distributionscenter

Kör: python merge_datasets.py
"""

import json
import math
import logging
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-8s %(message)s")

HQ_FILE       = "routerider_foretag.xlsx"
LAGER_FILE    = "routerider_lager.json"
OUTPUT_FILE   = "routerider_komplett.xlsx"

# Avstånd (meter) inom vilket en lagerpunkt och ett HQ anses vara samma plats.
# Om lager är > SAME_SITE_THRESHOLD m från HQ → separat rad (eget upphämtningsstopp).
SAME_SITE_THRESHOLD = 500


def haversine_m(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    """Avstånd i meter mellan två koordinater."""
    R = 6_371_000
    p = math.pi / 180
    a = (
        math.sin((lat2 - lat1) * p / 2) ** 2
        + math.cos(lat1 * p) * math.cos(lat2 * p) * math.sin((lng2 - lng1) * p / 2) ** 2
    )
    return 2 * R * math.asin(math.sqrt(a))


def load_hq(path: str) -> List[Dict]:
    df = pd.read_excel(path, sheet_name="Företag", dtype=str)
    df = df.fillna("")
    records = df.to_dict("records")
    for r in records:
        r["pickup_type"] = "HQ"
        # Konvertera koordinater till float om de finns
        try:
            r["lat_f"] = float(r.get("Lat", "") or 0) or None
            r["lng_f"] = float(r.get("Lng", "") or 0) or None
        except ValueError:
            r["lat_f"] = r["lng_f"] = None
    return records


def load_lager(path: str) -> List[Dict]:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    for item in data:
        item["pickup_type"] = item.get("typ", "Lager").capitalize()
        item["lat_f"] = item.get("lat")
        item["lng_f"] = item.get("lng")
    return data


def match_and_merge(hq_list: List[Dict], lager_list: List[Dict]) -> List[Dict]:
    """
    Logik:
    - Om ett lager-objekt matchar ett HQ på org.nr → lägg till pickup_type = "Lager (filial till HQ)"
      och markera HQ-raden att den har ett externt lager.
    - Om lagret är < SAME_SITE_THRESHOLD m från HQ-adressen → flagga HQ-raden, skippa lager-raden.
    - Annars → lägg till lager-objektet som en separat rad.
    """
    merged: List[Dict] = list(hq_list)
    hq_orgnr_map = {r.get("Org.nr", ""): r for r in hq_list if r.get("Org.nr")}

    for lager in lager_list:
        org_nr  = lager.get("org_nr", "")
        matching_hq = hq_orgnr_map.get(org_nr)

        if matching_hq:
            # Kolla om lagret är på samma adress som HQ
            hlat, hlng = matching_hq.get("lat_f"), matching_hq.get("lng_f")
            llat, llng = lager.get("lat_f"), lager.get("lng_f")

            if hlat and llat:
                dist = haversine_m(hlat, hlng, llat, llng)
                if dist < SAME_SITE_THRESHOLD:
                    # Samma plats → bara flagga HQ-raden
                    matching_hq["Anteckningar"] = (
                        matching_hq.get("Anteckningar", "") +
                        f" [Lager bekräftat på samma adress, källa: {lager['source']}]"
                    ).strip()
                    continue

            # Lager på annan adress → eget upphämtningsstopp
            new_row = _lager_to_row(lager, parent_hq=matching_hq)
            merged.append(new_row)
            matching_hq["Anteckningar"] = (
                matching_hq.get("Anteckningar", "") +
                f" [Har externt lager i {lager.get('city', '')}]"
            ).strip()
        else:
            # Inget matchande HQ — lägg till som fristående lager-rad
            merged.append(_lager_to_row(lager, parent_hq=None))

    return merged


def _lager_to_row(lager: Dict, parent_hq: Optional[Dict]) -> Dict:
    """Konverterar ett lager-objekt till samma kolumnformat som HQ-raden."""
    omsattning = 0
    if parent_hq:
        try:
            omsattning = int(str(parent_hq.get("Omsättning (kr)", "0")).replace(" ", "") or 0)
        except ValueError:
            omsattning = 0

    score = lager.get("godspotential") or _estimate_score(omsattning, lager.get("typ", ""))

    return {
        "Företagsnamn":      lager.get("name", ""),
        "Org.nr":            lager.get("org_nr", ""),
        "Stad (korridor)":   lager.get("city", ""),
        "Gatuadress":        (lager.get("adress") or "").strip(),
        "Postnr":            lager.get("postnr", ""),
        "Ort":               lager.get("ort", ""),
        "Lat":               lager.get("lat"),
        "Lng":               lager.get("lng"),
        "Bransch":           lager.get("bransch", ""),
        "SNI-kod":           "",
        "Omsättning (kr)":   omsattning or (parent_hq or {}).get("Omsättning (kr)", ""),
        "Anställda":         "",
        "Hemsida":           lager.get("hemsida", ""),
        "Godspotential/mån": score,
        "Prioritet":         _prioritet(score),
        "Outreach-status":   "Ej kontaktad",
        "Kontaktperson":     "",
        "Telefon":           lager.get("telefon", ""),
        "Email":             "",
        "Anteckningar":      f"Källa: {lager.get('source', '')} | Typ: {lager.get('typ', '')}",
        "Allabolag-URL":     lager.get("allabolag_url", ""),
        "pickup_type":       lager.get("pickup_type", "Lager"),
        "lat_f":             lager.get("lat"),
        "lng_f":             lager.get("lng"),
    }


def _estimate_score(omsattning_kr: int, typ: str) -> int:
    typ = typ.lower()
    if "distribut" in typ or "terminal" in typ:
        faktor = 3.5
    elif "lager" in typ:
        faktor = 3.0
    elif "fabrik" in typ:
        faktor = 2.5
    else:
        faktor = 1.5
    return max(1, int((omsattning_kr / 1_000_000) * faktor))


def _prioritet(score: int) -> str:
    if score >= 50: return "H"
    if score >= 15: return "M"
    return "L"


# ─── Excel-export ─────────────────────────────────────────────────────────────

PICKUP_COLORS = {
    "HQ":                   "DDEEFF",
    "Lager":                "E2EFDA",
    "Lager (filial till HQ)":"C6EFCE",
    "Distributionscenter":  "FFF2CC",
    "Fabrik":               "FCE4D6",
    "Industrilokal":        "EDEDED",
    "Industri":             "EDEDED",
}

PRIORITY_COLORS = {"H": "00B050", "M": "FFC000", "L": "FF0000"}

OUTPUT_COLUMNS = [
    "pickup_type", "Prioritet", "Godspotential/mån",
    "Företagsnamn", "Org.nr", "Stad (korridor)",
    "Gatuadress", "Postnr", "Ort", "Lat", "Lng",
    "Bransch", "SNI-kod", "Omsättning (kr)", "Anställda",
    "Hemsida", "Outreach-status",
    "Kontaktperson", "Telefon", "Email",
    "Anteckningar", "Allabolag-URL",
]

RENAME = {"pickup_type": "Upphämtningstyp"}


def export(rows: List[Dict], path: str = OUTPUT_FILE) -> None:
    # Normalisera rader till OUTPUT_COLUMNS
    normalized = []
    for r in rows:
        normalized.append({col: r.get(col, "") for col in OUTPUT_COLUMNS})

    df = pd.DataFrame(normalized, columns=OUTPUT_COLUMNS)
    df.rename(columns=RENAME, inplace=True)
    df.sort_values(["Prioritet", "Godspotential/mån"], ascending=[True, False], inplace=True)

    prio_col_name = "Prioritet"
    type_col_name = "Upphämtningstyp"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Alla upphämtningspunkter")
        ws = writer.sheets["Alla upphämtningspunkter"]

        # Rubrikrad
        hfill = PatternFill("solid", fgColor="0D3057")
        hfont = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 24

        # Identifiera kolumnindex
        headers = [cell.value for cell in ws[1]]
        prio_idx = headers.index(prio_col_name) if prio_col_name in headers else None
        type_idx = headers.index(type_col_name) if type_col_name in headers else None

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            prio = row[prio_idx].value if prio_idx is not None else ""
            typ  = row[type_idx].value if type_idx is not None else ""

            bg = PICKUP_COLORS.get(typ, "FFFFFF")
            fill = PatternFill("solid", fgColor=bg)
            for cell in row:
                cell.fill = fill

            # Färgkoda prioritet-cellen
            if prio_idx is not None:
                p_cell = row[prio_idx]
                p_cell.font = Font(bold=True, color=PRIORITY_COLORS.get(prio, "000000"))

        # Kolumnbredder
        widths = {
            "Upphämtningstyp": 22, "Företagsnamn": 38, "Gatuadress": 30,
            "Bransch": 32, "Hemsida": 32, "Allabolag-URL": 40,
            "Omsättning (kr)": 18, "Anteckningar": 35,
        }
        for i, header in enumerate(headers, 1):
            col_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[col_letter].width = widths.get(header, 13)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    log.info(f"Exporterade {len(rows)} rader → {path}")


# ─── Huvudprogram ──────────────────────────────────────────────────────────────

def main() -> None:
    if not Path(HQ_FILE).exists():
        log.error(f"{HQ_FILE} saknas — kör scraper.py först")
        return
    if not Path(LAGER_FILE).exists():
        log.error(f"{LAGER_FILE} saknas — kör warehouse_finder.py först")
        return

    hq_list    = load_hq(HQ_FILE)
    lager_list = load_lager(LAGER_FILE)

    log.info(f"HQ-poster:    {len(hq_list)}")
    log.info(f"Lager-poster: {len(lager_list)}")

    merged = match_and_merge(hq_list, lager_list)

    # Räkna typer
    type_counts: Dict[str, int] = {}
    for r in merged:
        t = r.get("pickup_type", "?")
        type_counts[t] = type_counts.get(t, 0) + 1

    log.info(f"\nResultat:")
    for t, n in sorted(type_counts.items()):
        log.info(f"  {t:<30} {n:>4} st")
    log.info(f"  {'TOTALT':<30} {len(merged):>4} upphämtningspunkter")

    export(merged)
    print(f"\n✓ Klar! Öppna {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
