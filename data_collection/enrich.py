#!/usr/bin/env python3
"""
RouteRider — Enrichment
Fyller på routerider_lager.json med:
  1. Gatuadress (Nominatim reverse geocoding)
  2. org_nr, omsättning, bransch (allabolag.se via Playwright)
  3. Exporterar routerider_lager_enriched.xlsx
"""

import asyncio
import json
import logging
import re
import time
import urllib.parse
from pathlib import Path

import pandas as pd
from bs4 import BeautifulSoup
from geopy.extra.rate_limiter import RateLimiter
from geopy.geocoders import Nominatim
from openpyxl.styles import Font, PatternFill
from playwright.async_api import async_playwright

log = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s %(message)s",
)

INPUT_JSON   = Path("routerider_lager.json")
OUTPUT_JSON  = Path("routerider_lager_enriched.json")
OUTPUT_XLSX  = Path("routerider_lager_enriched.xlsx")
CHECKPOINT   = Path("enrich_checkpoint.json")

BRANSCH_FAKTOR = {
    "46": 2.8,
    "52": 3.5,
    "10": 3.0,
    "20": 2.5,
    "25": 2.3,
    "28": 2.2,
    "29": 2.0,
    "45": 1.5,
    "47": 1.2,
}

def _sni_prefix(bransch: str) -> str:
    m = re.search(r"\d+", bransch or "")
    return m.group()[:2] if m else ""

def godspotential(omsattning_kr: float, sni_prefix: str) -> float:
    faktor = BRANSCH_FAKTOR.get(sni_prefix, 1.0)
    return round((omsattning_kr / 1_000_000) * faktor, 1)

# ── 1. Reverse geocoding ─────────────────────────────────────────────────────

def reverse_geocode_all(data: list) -> list:
    geolocator = Nominatim(user_agent="RouteRider/1.0")
    reverse = RateLimiter(geolocator.reverse, min_delay_seconds=1.1)

    need = [i for i, d in enumerate(data)
            if not (d.get("adress", "").strip() and d.get("postnr", ""))]
    log.info(f"  Reverse geocodar {len(need)} av {len(data)} poster...")

    for count, i in enumerate(need):
        d = data[i]
        lat, lng = d.get("lat"), d.get("lng")
        if not lat or not lng:
            continue
        try:
            loc = reverse(f"{lat}, {lng}", language="sv")
            if loc:
                raw = loc.raw.get("address", {})
                road   = raw.get("road", "")
                number = raw.get("house_number", "")
                postcode = raw.get("postcode", "").replace(" ", "")
                city = (raw.get("city") or raw.get("town")
                        or raw.get("village") or raw.get("municipality") or "")
                if road:
                    data[i]["adress"] = f"{road} {number}".strip()
                if postcode:
                    data[i]["postnr"] = postcode
                if city and not data[i].get("ort"):
                    data[i]["ort"] = city
        except Exception as e:
            log.warning(f"Geocoding fel ({d.get('name','?')}): {e}")

        if (count + 1) % 100 == 0:
            log.info(f"    Geocodning: {count+1}/{len(need)}")

    return data

# ── 2. Allabolag.se lookup ───────────────────────────────────────────────────

async def _accept_cookies(page):
    try:
        btn = page.locator(".qc-cmp2-summary-buttons button:last-child")
        if await btn.count():
            await btn.first.click()
            await page.wait_for_timeout(600)
    except Exception:
        pass

async def lookup_allabolag(page, name: str, city: str) -> dict:
    result: dict = {}
    try:
        q   = urllib.parse.quote(name)
        loc = urllib.parse.quote(city)
        url = f"https://www.allabolag.se/vad/{q}/var/{loc}"
        await page.goto(url, wait_until="domcontentloaded", timeout=25_000)
        await page.wait_for_timeout(1_200)
        await _accept_cookies(page)

        # Find first link to a company page (/foretag/)
        links = page.locator("a[href*='/foretag/']")
        if not await links.count():
            return result

        href = await links.first.get_attribute("href")
        if not href:
            return result

        company_url = (
            f"https://www.allabolag.se{href}"
            if href.startswith("/") else href
        )
        await page.goto(company_url, wait_until="domcontentloaded", timeout=25_000)
        await page.wait_for_timeout(1_200)
        await _accept_cookies(page)

        soup = BeautifulSoup(await page.content(), "lxml")

        # org_nr — 6 siffror-4 siffror mönster
        for text in soup.find_all(string=True):
            m = re.search(r"\b(\d{6}-\d{4})\b", text)
            if m:
                result["org_nr"] = m.group(1)
                break

        # Bransch / SNI — letaar efter "SNI" i texten
        for el in soup.find_all(["dt", "td", "th", "span", "li"]):
            txt = el.get_text(" ", strip=True)
            if "SNI" in txt or "Bransch" in txt:
                sib = el.find_next_sibling()
                if sib:
                    result["bransch"] = sib.get_text(" ", strip=True)[:120]
                    break

        # Omsättning / nettoomsättning
        for el in soup.find_all(["td", "dd", "span", "div"]):
            txt = el.get_text(strip=True)
            if re.search(r"(?i)(netto)?omsättning", txt):
                sib = el.find_next_sibling()
                candidate = sib.get_text(strip=True) if sib else ""
                m = re.search(r"[\d\s]+", candidate)
                if m:
                    val = m.group().replace(" ", "").strip()
                    if val.isdigit() and len(val) >= 3:
                        # Values on allabolag are in kkr (thousands of SEK)
                        result["omsattning_kr"] = int(val) * 1_000
                        break

        result["allabolag_url"] = company_url

    except Exception as e:
        log.debug(f"Allabolag lookup '{name}': {e}")

    return result


async def enrich_named(data: list) -> list:
    # Load checkpoint
    done: set[str] = set()
    if CHECKPOINT.exists():
        with open(CHECKPOINT) as f:
            done = set(json.load(f).get("done", []))
        log.info(f"  Checkpoint: {len(done)} företag redan klara")

    named_idx = [i for i, d in enumerate(data) if d.get("name", "").strip()]
    todo = [i for i in named_idx if data[i]["name"] not in done]
    log.info(f"  {len(todo)} företag att berika (av {len(named_idx)} namngivna)...")

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            locale="sv-SE",
        )
        page = await ctx.new_page()

        for seq, i in enumerate(todo, 1):
            d = data[i]
            name = d["name"].strip()
            city = d.get("city") or d.get("ort") or ""

            log.info(f"  [{seq}/{len(todo)}] {name} ({city})")

            enriched = await lookup_allabolag(page, name, city)
            if enriched:
                data[i].update(enriched)
                sni = _sni_prefix(enriched.get("bransch", ""))
                data[i]["godspotential"] = godspotential(
                    enriched.get("omsattning_kr", 0), sni
                )
                log.info(
                    f"    ✓ org={enriched.get('org_nr','–')} "
                    f"omsättn={enriched.get('omsattning_kr',0):,} kr"
                )
            else:
                log.info("    – ej hittad")

            done.add(name)

            if seq % 10 == 0:
                _save_checkpoint(done, data)

            await asyncio.sleep(2.0)

        await browser.close()

    _save_checkpoint(done, data)
    return data


def _save_checkpoint(done: set, data: list):
    with open(CHECKPOINT, "w") as f:
        json.dump({"done": list(done)}, f)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ── 3. Excel export ──────────────────────────────────────────────────────────

TYP_COLORS = {
    "lager":              "D6EAF8",
    "industri":           "D5F5E3",
    "industrilokal":      "FDEBD0",
    "fabrik":             "F9EBEA",
    "distributionscenter":"EBF5FB",
}

def export_excel(data: list):
    rows = []
    for d in data:
        name = d.get("name", "").strip()
        adress = d.get("adress", "").strip()
        if not name and not adress:
            continue  # Skip completely anonymous entries

        sni = _sni_prefix(d.get("bransch", ""))
        omsattn = d.get("omsattning_kr", 0)
        gp = d.get("godspotential") or godspotential(omsattn, sni)

        rows.append({
            "Namn":             name or "–",
            "Typ":              d.get("typ", ""),
            "Stad":             d.get("city") or d.get("ort") or "",
            "Adress":           adress,
            "Postnr":           d.get("postnr", ""),
            "Lat":              d.get("lat"),
            "Lng":              d.get("lng"),
            "Org.nr":           d.get("org_nr", ""),
            "Bransch":          d.get("bransch", ""),
            "SNI":              sni,
            "Omsättning (kr)":  omsattn,
            "Godspotential":    gp,
            "Hemsida":          d.get("hemsida", ""),
            "Telefon":          d.get("telefon", ""),
            "Allabolag":        d.get("allabolag_url", ""),
            "Källa":            d.get("source", ""),
        })

    df = pd.DataFrame(rows)
    df = df.sort_values(["Godspotential", "Omsättning (kr)"], ascending=False)

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Lokaler")
        ws = writer.sheets["Lokaler"]

        # Header style
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2C3E50")

        # Row colors by typ
        typ_col = df.columns.get_loc("Typ")
        for row_idx, typ_val in enumerate(df["Typ"], start=2):
            color = TYP_COLORS.get(str(typ_val).lower(), "FFFFFF")
            fill  = PatternFill("solid", fgColor=color)
            for cell in ws[row_idx]:
                cell.fill = fill

        # Column widths
        for col in ws.columns:
            vals = [str(c.value) for c in col if c.value]
            width = min(max((len(v) for v in vals), default=10) + 2, 55)
            ws.column_dimensions[col[0].column_letter].width = width

    log.info(f"Excel sparad: {OUTPUT_XLSX}  ({len(rows)} rader)")

# ── Main ─────────────────────────────────────────────────────────────────────

async def main():
    log.info("Laddar routerider_lager.json...")
    with open(INPUT_JSON, encoding="utf-8") as f:
        data = json.load(f)
    log.info(f"  {len(data)} lokaler")

    log.info("Steg 1 — Reverse geocoding...")
    data = reverse_geocode_all(data)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    log.info("Steg 2 — Allabolag enrichment...")
    data = await enrich_named(data)

    log.info("Steg 3 — Exporterar Excel...")
    export_excel(data)

    named_with_org = sum(1 for d in data if d.get("org_nr"))
    log.info(
        f"\n✓ Klar!  {named_with_org} företag med org_nr  "
        f"| {len(data)} lokaler totalt"
    )
    log.info(f"  {OUTPUT_XLSX}")
    log.info(f"  {OUTPUT_JSON}")


if __name__ == "__main__":
    asyncio.run(main())
