#!/usr/bin/env python3
"""
RouteRider — Steg 1: Datainsamling
Scrapar Allabolag.se med Playwright (headless Chromium) för att rendera
JavaScript och fånga nätverksanrop som ger oss full företagsdata.

Kör: python scraper.py
Kräver: pip install -r requirements.txt && playwright install chromium
"""

import asyncio
import json
import logging
import re
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.async_api import async_playwright, Page, Route

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════════════════
# KONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

# Allabolag intern plats-ID (hämtas via /api/locations)
CITIES: List[Dict] = [
    # ── Norra grenen (E18/E20): Stockholm → Örebro ───────────────────────────
    {"name": "Enköping",    "id": 176},
    {"name": "Västerås",    "id": 235},
    {"name": "Eskilstuna",  "id": 166},
    {"name": "Strängnäs",   "id": 172},
    {"name": "Katrineholm", "id": 169},
    {"name": "Hallsberg",   "id": 289},
    {"name": "Örebro",      "id": 298},
    # ── Södra grenen (E4): Stockholm → Jönköping ─────────────────────────────
    {"name": "Södertälje",  "id": 156},
    {"name": "Flen",        "id": 167},
    {"name": "Nyköping",    "id": 170},
    {"name": "Norrköping",  "id": 306},
    {"name": "Linköping",   "id": 303},
    {"name": "Jönköping",   "id": 59},
    # ── Gemensam sträcka → Göteborg ──────────────────────────────────────────
    {"name": "Skövde",      "id": 267},
    {"name": "Falköping",   "id": 244},
    {"name": "Ulricehamn",  "id": 280},
    {"name": "Borås",       "id": 241},
    {"name": "Alingsås",    "id": 238},
]

# Allabolag industry-ID:n att söka på
# (deras eget system — inte SNI-koder)
# Dessa är industrikategorier i deras "supplier listing"
INDUSTRY_IDS: List[int] = list(range(1, 200))  # Breda sökning

# Alternativ: sök via SNI-koder direkt i URL-sökning
# SNI 10-33 = Tillverkning, 45-47 = Handel, 52 = Lager
TARGET_SNI_PREFIXES: List[str] = [
    "10", "11", "13", "15", "16", "17", "20", "21", "22", "23",
    "24", "25", "26", "27", "28", "29", "30", "31", "32",
    "45", "46", "47", "52",
]

MIN_OMSATTNING_KR = 5_000_000
OUTPUT_FILE       = "routerider_foretag.xlsx"
CHECKPOINT_FILE   = "checkpoint.json"

# ═══════════════════════════════════════════════════════════════════════════════
# ALLABOLAG API-KLIENTEN (fångar nätverksanrop via Playwright)
# ═══════════════════════════════════════════════════════════════════════════════

BASE_URL = "https://www.allabolag.se"

async def fetch_via_api(page: Page, location_id: int, industry_id: int, city_name: str) -> List[Dict]:
    """
    Öppnar Allabolag i browser, fångar det interna API-anropet och
    returnerar listan med företag (samma format som /api/search returnerar).
    """
    captured: List[Dict] = []

    async def intercept(route: Route, _):
        # Skicka requesten vidare och fånga svaret
        response = await route.fetch()
        try:
            body = await response.json()
            if isinstance(body, dict) and "companies" in body:
                companies = body.get("companies", [])
                if companies:
                    captured.extend(companies)
                    log.debug(f"    Fångade {len(companies)} företag via API-intercept")
        except Exception:
            pass
        await route.fulfill(response=response)

    # Lyssna på API-anrop mot /api/search
    await page.route("**/api/search**", intercept)

    try:
        url = f"{BASE_URL}/bransch-s%C3%B6k?locationId={location_id}&industryId={industry_id}"
        await page.goto(url, wait_until="networkidle", timeout=30_000)
        # Extra väntan för lazy-loading
        await page.wait_for_timeout(2000)
    except Exception as e:
        log.warning(f"Sida laddades inte: {e}")
    finally:
        await page.unroute("**/api/search**")

    return captured


async def fetch_direct_api(location_id: int, industry_id: int) -> List[Dict]:
    """
    Direkt API-anrop (fungerar för supplier-listing — betalade annonsörer).
    Används som komplement till browser-scraping.
    """
    import aiohttp
    url = f"{BASE_URL}/api/search?locationId={location_id}&industryId={industry_id}&size=100"
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
        "Accept": "application/json",
        "Referer": BASE_URL,
    }
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=15)) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    return data.get("companies", [])
    except Exception:
        pass
    return []


# ═══════════════════════════════════════════════════════════════════════════════
# SNI-SÖKNING VIA ALLABOLAG BOLAGSSIDOR
# ═══════════════════════════════════════════════════════════════════════════════

async def search_by_sni_and_city(page: Page, sni: str, city_name: str) -> List[Dict]:
    """
    Söker Allabolag efter företag med ett SNI-prefix i en given stad.
    Navigerar till söksidan och extraherar org.nr från resultaten.
    """
    companies = []
    captured_responses = []

    async def capture(route: Route, _):
        response = await route.fetch()
        try:
            body = await response.json()
            if isinstance(body, dict):
                # Fånga alla API-svar som innehåller org.nr / company-data
                hits = body.get("companies") or body.get("hits") or body.get("results") or []
                if hits and isinstance(hits, list):
                    captured_responses.extend(hits)
        except Exception:
            pass
        await route.fulfill(response=response)

    await page.route("**", capture)

    try:
        # Sök efter SNI + stad i Allabolags sökruta
        search_url = f"{BASE_URL}/search?query={sni}+{city_name}"
        await page.goto(search_url, wait_until="networkidle", timeout=30_000)
        await page.wait_for_timeout(3000)

        # Extrahera org.nr från HTML (som fallback)
        content = await page.content()
        orgnr_matches = re.findall(r"/(\d{6}-\d{4})", content)
        for orgnr in set(orgnr_matches):
            companies.append({"orgnr": orgnr.replace("-", ""), "sni_hint": sni, "city_hint": city_name})

    except Exception as e:
        log.warning(f"SNI-sökning misslyckades ({sni}, {city_name}): {e}")
    finally:
        await page.unroute("**")

    companies.extend(captured_responses)
    return companies


# ═══════════════════════════════════════════════════════════════════════════════
# HÄMTA DETALJDATA FÖR ETT BOLAG
# ═══════════════════════════════════════════════════════════════════════════════

async def get_company_details(page: Page, orgnr: str) -> Optional[Dict]:
    """
    Hämtar detaljerad data för ett specifikt bolag via dess Allabolag-sida.
    Fångar API-anropet som sidan gör för att hämta bolagsdata.
    """
    # Normalisera org.nr-format: 5568123456 -> 556812-3456
    if len(orgnr) == 10 and "-" not in orgnr:
        orgnr_fmt = orgnr[:6] + "-" + orgnr[6:]
    else:
        orgnr_fmt = orgnr

    captured = {}

    async def capture(route: Route, _):
        response = await route.fetch()
        try:
            body = await response.json()
            if isinstance(body, dict) and any(k in body for k in ("name","legalName","companyId","orgnr")):
                captured.update(body)
        except Exception:
            pass
        await route.fulfill(response=response)

    await page.route(f"**/{orgnr_fmt.replace('-','')}**", capture)
    await page.route("**/api/company/**", capture)

    try:
        url = f"{BASE_URL}/{orgnr_fmt}"
        await page.goto(url, wait_until="networkidle", timeout=25_000)
        await page.wait_for_timeout(1500)

        # Fallback: extrahera från HTML om API inte fångades
        if not captured:
            content = await page.content()
            captured = extract_from_html(content, orgnr_fmt)

    except Exception as e:
        log.warning(f"Kunde inte hämta detaljer för {orgnr_fmt}: {e}")
    finally:
        await page.unroute(f"**/{orgnr_fmt.replace('-','')}**")
        await page.unroute("**/api/company/**")

    return normalize_company(captured, orgnr_fmt) if captured else None


def extract_from_html(html: str, orgnr: str) -> Dict:
    """Extraherar bolagsdata från HTML när API-intercept inte funkar."""
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "lxml")
    c: Dict = {"orgnr": orgnr.replace("-", "")}
    full_text = soup.get_text(" ", strip=True)

    # Namn
    for sel in ["h1", "title"]:
        el = soup.select_one(sel)
        if el:
            c["name"] = el.get_text(strip=True).split("|")[0].strip()
            break

    # Adress
    for sel in ["[itemprop='streetAddress']", ".address"]:
        el = soup.select_one(sel)
        if el:
            c["visitorAddress"] = {"addressLine": el.get_text(strip=True)}
            break

    # Omsättning
    m = re.search(r"(\d[\d\s]+)\s*(?:MSEK|TSEK|MKR|TKR|kr)", full_text, re.I)
    if m:
        c["revenue"] = m.group(1).replace(" ", "")

    # Telefon
    phone_m = re.search(r"0[\d\s-]{8,12}", full_text)
    if phone_m:
        c["phone"] = phone_m.group(0).strip()

    return c


def normalize_company(raw: Dict, orgnr: str) -> Dict:
    """Konverterar Allabolags API-format till vårt standardformat."""
    # Koordinater
    lat, lng = None, None
    coords = (raw.get("location") or {}).get("coordinates", [])
    if coords:
        lat = coords[0].get("ycoordinate")
        lng = coords[0].get("xcoordinate")

    # Adress
    addr = raw.get("visitorAddress") or raw.get("postalAddress") or {}
    adress_line = addr.get("addressLine", "")
    postnr      = addr.get("zipCode", "")
    ort         = addr.get("postPlace", "")

    # Omsättning
    rev_raw = raw.get("revenue", "") or ""
    try:
        omsattning_kr = int(str(rev_raw).replace(" ", "")) if rev_raw else 0
    except ValueError:
        omsattning_kr = 0

    # Anställda
    employees = raw.get("employees", "") or ""

    # Bransch (Allabolags eget system)
    industries = raw.get("industries", []) or []
    bransch = industries[0].get("name", "") if industries else ""

    # Kontaktperson
    contact = raw.get("contactPerson") or {}
    kontaktperson = contact.get("name", "")
    kontakt_roll  = contact.get("role", "")

    orgnr_clean = orgnr.replace("-", "")
    if len(orgnr_clean) == 10:
        orgnr_fmt = orgnr_clean[:6] + "-" + orgnr_clean[6:]
    else:
        orgnr_fmt = orgnr

    return {
        "namn":         raw.get("name") or raw.get("legalName") or "",
        "org_nr":       orgnr_fmt,
        "adress":       adress_line.strip(),
        "postnr":       (postnr or "").replace(" ", ""),
        "ort":          ort,
        "lat":          lat,
        "lng":          lng,
        "bransch":      bransch,
        "sni_kod":      "",   # Allabolag-API ger inte alltid SNI
        "omsattning_kr": omsattning_kr,
        "anstallda":    employees,
        "hemsida":      raw.get("homePage") or "",
        "telefon":      raw.get("phone") or "",
        "email":        raw.get("email") or "",
        "kontaktperson": f"{kontaktperson} ({kontakt_roll})".strip(" ()") if kontaktperson else "",
        "allabolag_url": f"{BASE_URL}/{orgnr_fmt}",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# GODSPOTENTIAL & PRIORITET
# ═══════════════════════════════════════════════════════════════════════════════

BRANSCH_FAKTOR: Dict[str, float] = {
    "10": 3.0,  "11": 2.5,  "13": 2.0,  "15": 2.2,
    "16": 1.8,  "17": 1.8,  "20": 2.2,  "21": 2.5,
    "22": 2.0,  "23": 2.0,  "24": 2.5,  "25": 2.0,
    "28": 2.2,  "29": 2.0,  "45": 1.5,  "46": 2.8,
    "47": 1.8,  "52": 3.5,
}

def godspotential(company: Dict) -> int:
    sni = (company.get("sni_kod") or "")[:2]
    faktor = BRANSCH_FAKTOR.get(sni, 1.2)
    msek = company.get("omsattning_kr", 0) / 1_000_000
    return max(1, int(msek * faktor))

def prioritet(score: int) -> str:
    if score >= 50: return "H"
    if score >= 15: return "M"
    return "L"


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL-EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

COLUMNS = [
    ("Företagsnamn",       "namn"),
    ("Org.nr",             "org_nr"),
    ("Stad (korridor)",    "stad"),
    ("Gatuadress",         "adress"),
    ("Postnr",             "postnr"),
    ("Ort",                "ort"),
    ("Lat",                "lat"),
    ("Lng",                "lng"),
    ("Bransch",            "bransch"),
    ("SNI-kod",            "sni_kod"),
    ("Omsättning (kr)",    "omsattning_kr"),
    ("Anställda",          "anstallda"),
    ("Hemsida",            "hemsida"),
    ("Telefon",            "telefon"),
    ("Email",              "email"),
    ("Kontaktperson",      "kontaktperson"),
    ("Godspotential/mån",  "godspotential"),
    ("Prioritet",          "prioritet"),
    ("Outreach-status",    "outreach_status"),
    ("Anteckningar",       "anteckningar"),
    ("Allabolag-URL",      "allabolag_url"),
]

PRIORITY_COLORS = {"H": "C6EFCE", "M": "FFEB9C", "L": "FFC7CE"}
COL_WIDTHS = {
    "Företagsnamn": 38, "Gatuadress": 32, "Bransch": 35,
    "Hemsida": 35, "Allabolag-URL": 42, "Omsättning (kr)": 20,
    "Kontaktperson": 28, "Anteckningar": 30,
}

def export_excel(companies: List[Dict], path: str = OUTPUT_FILE) -> str:
    rows = []
    for c in companies:
        c["godspotential"]   = godspotential(c)
        c["prioritet"]       = prioritet(c["godspotential"])
        c["outreach_status"] = "Ej kontaktad"
        c.setdefault("anteckningar", "")
        rows.append({label: c.get(key, "") for label, key in COLUMNS})

    df = pd.DataFrame(rows, columns=[col for col, _ in COLUMNS])
    df.sort_values(["Prioritet", "Godspotential/mån"], ascending=[True, False], inplace=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Företag")
        ws = writer.sheets["Företag"]

        fill_h = PatternFill("solid", fgColor="1F3A6E")
        font_h = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws[1]:
            cell.fill = fill_h
            cell.font = font_h
            cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        prio_col = next(i for i, (l, _) in enumerate(COLUMNS, 1) if l == "Prioritet")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            prio = row[prio_col - 1].value or ""
            bg = PRIORITY_COLORS.get(prio, "FFFFFF")
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=bg)

        for i, (label, _) in enumerate(COLUMNS, 1):
            col_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[col_letter].width = COL_WIDTHS.get(label, 14)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    log.info(f"Excel sparad: {path}  ({len(rows)} företag)")
    return path


# ═══════════════════════════════════════════════════════════════════════════════
# CHECKPOINT
# ═══════════════════════════════════════════════════════════════════════════════

def load_checkpoint() -> Dict:
    if Path(CHECKPOINT_FILE).exists():
        with open(CHECKPOINT_FILE) as f:
            data = json.load(f)
        log.info(f"Återupptar checkpoint — {len(data['companies'])} bolag redan insamlade")
        return data
    return {"companies": {}, "done_keys": []}

def save_checkpoint(companies: Dict, done: List[str]) -> None:
    with open(CHECKPOINT_FILE, "w") as f:
        json.dump({"companies": companies, "done_keys": done}, f, ensure_ascii=False, indent=2)


# ═══════════════════════════════════════════════════════════════════════════════
# HUVUDPROGRAM
# ═══════════════════════════════════════════════════════════════════════════════

async def run():
    checkpoint = load_checkpoint()
    companies  = checkpoint["companies"]
    done       = set(checkpoint["done_keys"])

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        context = await browser.new_context(
            locale="sv-SE",
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        )
        page = await context.new_page()

        # ── Fas A: direkt API-sökning (supplier listing) ─────────────────────
        log.info("Fas A: Hämtar supplier-listings via direkt API")
        for city in CITIES:
            for ind_id in range(1, 50):  # Prova industri-ID:n 1-49
                key = f"api|{city['id']}|{ind_id}"
                if key in done:
                    continue

                results = await fetch_direct_api(city["id"], ind_id)
                for raw in results:
                    orgnr = str(raw.get("orgnr", ""))
                    if not orgnr or orgnr in companies:
                        continue
                    # Normalisera och filtrera
                    c = normalize_company(raw, orgnr)
                    c["stad"] = city["name"]
                    if c.get("omsattning_kr", 0) >= MIN_OMSATTNING_KR:
                        companies[orgnr] = c
                        log.info(f"  + {c['namn']:<40} {c.get('omsattning_kr',0):>12,} kr  [{city['name']}]")

                done.add(key)
                if len(results) > 0:
                    save_checkpoint(companies, list(done))
                await asyncio.sleep(0.3)

        # ── Fas B: browser-sökning via SNI + stad ────────────────────────────
        log.info(f"\nFas B: Browser-sökning — {len(CITIES)} städer × {len(TARGET_SNI_PREFIXES)} SNI:er")
        for city in CITIES:
            for sni in TARGET_SNI_PREFIXES:
                key = f"sni|{city['name']}|{sni}"
                if key in done:
                    continue

                log.info(f"  {city['name']} SNI={sni}")
                raw_list = await search_by_sni_and_city(page, sni, city["name"])

                new_orgnrs = []
                for raw in raw_list:
                    orgnr = str(raw.get("orgnr","")).replace("-","")
                    if orgnr and len(orgnr) == 10 and orgnr not in companies:
                        new_orgnrs.append(orgnr)

                log.info(f"    → {len(new_orgnrs)} nya org.nr att hämta detaljer för")

                for orgnr in new_orgnrs:
                    c = await get_company_details(page, orgnr)
                    if c and c.get("omsattning_kr", 0) >= MIN_OMSATTNING_KR:
                        c["stad"] = city["name"]
                        companies[orgnr] = c
                        log.info(f"    + {c.get('namn','?'):<40} {c.get('omsattning_kr',0):>12,} kr")
                    await asyncio.sleep(1.0)

                done.add(key)
                save_checkpoint(companies, list(done))
                await asyncio.sleep(1.5)

        await browser.close()

    company_list = list(companies.values())
    log.info(f"\nTotalt: {len(company_list)} unika bolag")

    if company_list:
        export_excel(company_list)
        print(f"\n✓ Klar! Öppna {OUTPUT_FILE}")
        Path(CHECKPOINT_FILE).unlink(missing_ok=True)
    else:
        print("\n⚠ Inga bolag hittades. Se loggen för detaljer.")


def main():
    asyncio.run(run())

if __name__ == "__main__":
    main()
