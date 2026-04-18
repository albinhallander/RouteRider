#!/usr/bin/env python3
"""
RouteRider — Laddstationsfinnaren
Hittar laddstationer för tunga lastbilar längs Stockholm–Göteborg-korridoren.

Datakällor (i prioritetsordning):
  1. Overpass API (OpenStreetMap) — gratis, ingen nyckel
  2. Open Charge Map API          — gratis, ingen nyckel (global databas)
  3. NOBIL API                    — gratis, nordisk databas

Kör: python charging_station_finder.py
     python charging_station_finder.py --ocm-key <din_nyckel>  (valfritt, höjer rate-limit)

Output:
  routerider_laddstationer.json   — används av appen
  routerider_laddstationer.xlsx   — för manuell granskning
"""

import argparse
import json
import logging
import math
import time
from pathlib import Path
from typing import Dict, List, Optional

import requests

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-8s %(message)s")

# ═══════════════════════════════════════════════════════════════════════════════
# KONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

# Städer längs Stockholm–Göteborg-korridoren (samma som warehouse_finder)
CITIES_GEO: List[Dict] = [
    {"name": "Stockholm",    "lat": 59.3293, "lng": 18.0686, "radius_m": 20_000},
    {"name": "Södertälje",   "lat": 59.1960, "lng": 17.6260, "radius_m": 10_000},
    {"name": "Enköping",     "lat": 59.6350, "lng": 17.0770, "radius_m": 8_000},
    {"name": "Västerås",     "lat": 59.6100, "lng": 16.5450, "radius_m": 12_000},
    {"name": "Eskilstuna",   "lat": 59.3710, "lng": 16.5100, "radius_m": 10_000},
    {"name": "Strängnäs",    "lat": 59.3780, "lng": 17.0310, "radius_m": 7_000},
    {"name": "Katrineholm",  "lat": 58.9960, "lng": 16.2080, "radius_m": 7_000},
    {"name": "Hallsberg",    "lat": 59.0660, "lng": 15.1110, "radius_m": 6_000},
    {"name": "Örebro",       "lat": 59.2750, "lng": 15.2130, "radius_m": 12_000},
    {"name": "Flen",         "lat": 59.0580, "lng": 16.5880, "radius_m": 5_000},
    {"name": "Nyköping",     "lat": 58.7530, "lng": 17.0050, "radius_m": 8_000},
    {"name": "Norrköping",   "lat": 58.5940, "lng": 16.1880, "radius_m": 12_000},
    {"name": "Linköping",    "lat": 58.4110, "lng": 15.6210, "radius_m": 12_000},
    {"name": "Jönköping",    "lat": 57.7820, "lng": 14.1610, "radius_m": 12_000},
    {"name": "Skövde",       "lat": 58.3910, "lng": 13.8450, "radius_m": 8_000},
    {"name": "Falköping",    "lat": 58.1800, "lng": 13.5510, "radius_m": 7_000},
    {"name": "Ulricehamn",   "lat": 57.7930, "lng": 13.4200, "radius_m": 6_000},
    {"name": "Borås",        "lat": 57.7210, "lng": 12.9400, "radius_m": 10_000},
    {"name": "Alingsås",     "lat": 57.9300, "lng": 12.5330, "radius_m": 7_000},
    {"name": "Göteborg",     "lat": 57.7089, "lng": 11.9746, "radius_m": 20_000},
    # Tillägg längs E4/E20
    {"name": "Markaryd",     "lat": 56.3330, "lng": 13.5670, "radius_m": 6_000},
    {"name": "Ljungby",      "lat": 56.8500, "lng": 13.9330, "radius_m": 6_000},
    {"name": "Varberg",      "lat": 57.1000, "lng": 12.2330, "radius_m": 8_000},
    {"name": "Laholm",       "lat": 56.5500, "lng": 12.7670, "radius_m": 5_000},
    {"name": "Helsingborg",  "lat": 56.0461, "lng": 12.6941, "radius_m": 10_000},
]

HEADERS = {
    "User-Agent": "RouteRider/1.0 (laddstationsfinnare; kontakt: routerider@github.com)",
    "Accept-Language": "sv-SE,sv;q=0.9",
}

OUTPUT_JSON  = "routerider_laddstationer.json"
OUTPUT_EXCEL = "routerider_laddstationer.xlsx"

# ═══════════════════════════════════════════════════════════════════════════════
# KÄLLA 1: OVERPASS API (OpenStreetMap)
# ═══════════════════════════════════════════════════════════════════════════════

OVERPASS_URLS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
]

# Hämtar alla laddstationer — filtrerar för HGV-kompatibilitet efteråt
OSM_QUERY = """
[out:json][timeout:60];
(
  node["amenity"="charging_station"]{bbox};
  way["amenity"="charging_station"]{bbox};
);
out center tags;
"""


def overpass_query(city: Dict) -> List[Dict]:
    r    = city["radius_m"]
    lat  = city["lat"]
    lng  = city["lng"]
    dlat = r / 111_000
    dlng = r / (111_000 * math.cos(math.radians(lat)))
    bbox = f"({lat-dlat:.5f},{lng-dlng:.5f},{lat+dlat:.5f},{lng+dlng:.5f})"

    query = OSM_QUERY.replace("{bbox}", bbox)

    for attempt in range(3):
        if attempt:
            time.sleep(30 * attempt)
        for url in OVERPASS_URLS:
            try:
                resp = requests.post(
                    url, data={"data": query},
                    timeout=90, headers={"User-Agent": HEADERS["User-Agent"]}
                )
                resp.raise_for_status()
                data = resp.json()
                break
            except Exception as e:
                log.warning(f"Overpass fel ({url}, {city['name']}): {e}")
        else:
            continue
        break
    else:
        return []

    results = []
    for el in data.get("elements", []):
        tags = el.get("tags", {})
        if el["type"] == "node":
            elat, elng = el.get("lat"), el.get("lon")
        else:
            center = el.get("center", {})
            elat, elng = center.get("lat"), center.get("lon")
        if not elat:
            continue

        results.append({
            "source":          "openstreetmap",
            "city":            city["name"],
            "name":            tags.get("name") or tags.get("operator") or tags.get("brand") or "Laddstation",
            "lat":             round(elat, 6),
            "lng":             round(elng, 6),
            "address":         (tags.get("addr:street", "") + " " + tags.get("addr:housenumber", "")).strip(),
            "postcode":        tags.get("addr:postcode", ""),
            "operator":        tags.get("operator") or tags.get("brand") or "",
            "hgv_compatible":  _osm_hgv_compatible(tags),
            "charging_points": _osm_parse_int(tags.get("capacity")),
            "max_power_kw":    _osm_parse_float(tags.get("maxpower")),
            "connectors":      _osm_connectors(tags),
            "open_hours":      tags.get("opening_hours", ""),
            "status":          "LIVE",
            "osm_id":          el.get("id"),
        })

    log.info(f"  Overpass → {city['name']}: {len(results)} stationer")
    return results


def _osm_hgv_compatible(tags: Dict) -> bool:
    """Avgör om stationen är kompatibel med tunga fordon baserat på OSM-taggar."""
    # Explicit HGV-taggning
    if tags.get("hgv") == "yes":
        return True
    if tags.get("truck") == "yes":
        return True
    # Hög effekt indikerar lastbilsladdning (>= 150 kW)
    maxpower = _osm_parse_float(tags.get("maxpower") or tags.get("socket:type2_combo:output"))
    if maxpower and maxpower >= 150:
        return True
    # MCS-kontakt (Megawatt Charging System) = lastbil
    if "mcs" in str(tags).lower():
        return True
    # Operatörer som primärt riktar sig mot tunga fordon
    operator = (tags.get("operator") or tags.get("brand") or "").lower()
    truck_operators = ["einride", "kempower", "scania", "volvo trucks", "dhl charging", "vattenfall electric"]
    if any(op in operator for op in truck_operators):
        return True
    return False


def _osm_parse_int(val) -> Optional[int]:
    try:
        return int(str(val).strip())
    except (TypeError, ValueError):
        return None


def _osm_parse_float(val) -> Optional[float]:
    try:
        # Hantera "150 kW" → 150.0
        cleaned = str(val).lower().replace("kw", "").replace("w", "").strip()
        v = float(cleaned)
        # Om värde ser ut som watt (>= 10000) konvertera till kW
        return v / 1000 if v >= 10_000 else v
    except (TypeError, ValueError):
        return None


def _osm_connectors(tags: Dict) -> List[str]:
    connectors = []
    connector_map = {
        "socket:type2_combo": "CCS2",
        "socket:chademo": "CHAdeMO",
        "socket:type2": "Type2",
        "socket:tesla_supercharger": "Tesla",
        "socket:mcs": "MCS",
        "socket:ccs": "CCS",
    }
    for tag, name in connector_map.items():
        if tags.get(tag) or tags.get(f"{tag}:output"):
            connectors.append(name)
    return connectors


# ═══════════════════════════════════════════════════════════════════════════════
# KÄLLA 2: OPEN CHARGE MAP API
# ═══════════════════════════════════════════════════════════════════════════════

OCM_BASE = "https://api.openchargemap.io/v3/poi"

# Connection type IDs för snabbladdning (relevant för lastbilar)
# 25 = CCS Type 2 (Combined), 32 = CCS (Type 1 J1772), 2 = CHAdeMO
# 33 = Tesla (roadster), 27 = Tesla (Model S/X)
OCM_FAST_TYPES = {25, 32, 2, 30}  # 30 = Type 2 (AC, kan vara snabbt)


def ocm_query(city: Dict, api_key: Optional[str] = None) -> List[Dict]:
    params = {
        "latitude":         city["lat"],
        "longitude":        city["lng"],
        "distance":         city["radius_m"] / 1000,  # km
        "distanceunit":     "KM",
        "maxresults":       500,
        "compact":          True,
        "verbose":          False,
        "countrycode":      "SE",
        "levelid":          3,  # DC fast chargers only
    }
    if api_key:
        params["key"] = api_key

    try:
        resp = requests.get(OCM_BASE, params=params, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        log.error(f"OCM fel ({city['name']}): {e}")
        return []

    results = []
    for poi in data:
        addr_info = poi.get("AddressInfo", {})
        lat  = addr_info.get("Latitude")
        lng  = addr_info.get("Longitude")
        if not lat:
            continue

        connections = poi.get("Connections") or []
        max_kw      = max((c.get("PowerKW") or 0 for c in connections), default=None) or None
        conn_types  = list({
            c.get("ConnectionType", {}).get("Title", "") for c in connections
            if c.get("ConnectionType")
        })

        # HGV-kompatibel: >= 150 kW eller MCS
        hgv = bool(max_kw and max_kw >= 150) or any("mcs" in t.lower() for t in conn_types)

        status_type = poi.get("StatusType", {}) or {}
        is_live = status_type.get("IsOperational", True)

        results.append({
            "source":          "open_charge_map",
            "city":            city["name"],
            "name":            addr_info.get("Title") or "Laddstation",
            "lat":             round(lat, 6),
            "lng":             round(lng, 6),
            "address":         addr_info.get("AddressLine1") or "",
            "postcode":        addr_info.get("Postcode") or "",
            "operator":        (poi.get("OperatorInfo") or {}).get("Title") or "",
            "hgv_compatible":  hgv,
            "charging_points": poi.get("NumberOfPoints"),
            "max_power_kw":    max_kw,
            "connectors":      conn_types,
            "open_hours":      "",
            "status":          "LIVE" if is_live else "OFFLINE",
            "ocm_id":          poi.get("ID"),
        })

    log.info(f"  OCM → {city['name']}: {len(results)} stationer")
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# KÄLLA 3: NOBIL API (nordisk databas)
# ═══════════════════════════════════════════════════════════════════════════════

NOBIL_BASE = "https://nobil.no/api/server/datadump.php"


def nobil_query_sweden() -> List[Dict]:
    """
    Laddar ned hela Sverige-dumpen från NOBIL en gång (undviker per-stad-anrop).
    NOBIL ger en JSON-dump med alla svenska stationer.
    """
    params = {
        "apikey":      "demo",  # NOBIL tillåter demo-nyckel för läsning
        "countrycode": "SWE",
        "format":      "json",
        "type":        "json",
    }
    try:
        log.info("NOBIL: Laddar ned Sverige-dump...")
        resp = requests.get(NOBIL_BASE, params=params, headers=HEADERS, timeout=60)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        log.error(f"NOBIL fel: {e}")
        return []

    stations_raw = data.get("chargerstations", []) if isinstance(data, dict) else data
    results = []

    for raw in stations_raw:
        attrs = raw.get("attr", {})
        pos   = attrs.get("Position", {})
        lat   = _safe_float(pos.get("lat"))
        lng   = _safe_float(pos.get("lon") or pos.get("lng"))
        if not lat:
            continue

        # Kontaktorer och effekt
        conn_info = attrs.get("Connectors", {})
        connectors = []
        max_kw     = None
        for conn in conn_info.values() if isinstance(conn_info, dict) else []:
            ctype = conn.get("connectortype", {}).get("trans", "")
            if ctype:
                connectors.append(ctype)
            kw = _safe_float(conn.get("chargingcapacity"))
            if kw and (max_kw is None or kw > max_kw):
                max_kw = kw

        hgv = bool(max_kw and max_kw >= 150) or "truck" in str(attrs).lower()

        results.append({
            "source":          "nobil",
            "city":            attrs.get("Municipality", {}).get("trans", ""),
            "name":            attrs.get("name") or attrs.get("Street", {}).get("trans", "") or "Laddstation",
            "lat":             round(lat, 6),
            "lng":             round(lng, 6),
            "address":         attrs.get("Street", {}).get("trans", ""),
            "postcode":        attrs.get("Zipcode", {}).get("trans", ""),
            "operator":        attrs.get("Owned_by", {}).get("trans", ""),
            "hgv_compatible":  hgv,
            "charging_points": _safe_int(attrs.get("Number_charging_points")),
            "max_power_kw":    max_kw,
            "connectors":      connectors,
            "open_hours":      attrs.get("Open_Hours", {}).get("trans", ""),
            "status":          "LIVE",
            "nobil_id":        raw.get("id"),
        })

    log.info(f"NOBIL: {len(results)} stationer totalt i Sverige")
    return results


def _safe_float(val) -> Optional[float]:
    try:
        return float(str(val).replace(",", ".").strip())
    except (TypeError, ValueError):
        return None


def _safe_int(val) -> Optional[int]:
    try:
        return int(str(val).strip())
    except (TypeError, ValueError):
        return None


# ═══════════════════════════════════════════════════════════════════════════════
# DEDUPLICERING
# ═══════════════════════════════════════════════════════════════════════════════

SOURCE_PRIO = {"open_charge_map": 0, "nobil": 1, "openstreetmap": 2}


def deduplicate(stations: List[Dict], distance_m: float = 100) -> List[Dict]:
    """Slår ihop stationer som är < distance_m ifrån varandra."""

    def dist(a: Dict, b: Dict) -> float:
        if not (a.get("lat") and b.get("lat")):
            return float("inf")
        dlat = (a["lat"] - b["lat"]) * 111_000
        dlng = (a["lng"] - b["lng"]) * 111_000 * math.cos(math.radians(a["lat"]))
        return math.sqrt(dlat**2 + dlng**2)

    kept: List[Dict] = []
    for s in sorted(stations, key=lambda x: SOURCE_PRIO.get(x.get("source", ""), 9)):
        merged = False
        for existing in kept:
            if dist(s, existing) < distance_m:
                # Berika med data som saknas i den befintliga
                for key in ("operator", "charging_points", "max_power_kw", "open_hours", "address"):
                    if not existing.get(key) and s.get(key):
                        existing[key] = s[key]
                if not existing.get("hgv_compatible") and s.get("hgv_compatible"):
                    existing["hgv_compatible"] = True
                # Slå ihop connector-listor
                existing_conns = set(existing.get("connectors") or [])
                new_conns      = set(s.get("connectors") or [])
                existing["connectors"] = sorted(existing_conns | new_conns)
                existing["sources"] = list({existing.get("source", ""), s.get("source", "")})
                merged = True
                break
        if not merged:
            s.setdefault("sources", [s.get("source", "")])
            kept.append(s)

    return kept


def filter_corridor(stations: List[Dict]) -> List[Dict]:
    """Behåller bara stationer inom korridorens bounding box."""
    LAT_MIN, LAT_MAX = 56.0, 60.0
    LNG_MIN, LNG_MAX = 11.5, 18.5
    return [
        s for s in stations
        if s.get("lat") and s.get("lng")
        and LAT_MIN <= s["lat"] <= LAT_MAX
        and LNG_MIN <= s["lng"] <= LNG_MAX
    ]


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def export_json(stations: List[Dict], path: str = OUTPUT_JSON) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(stations, f, ensure_ascii=False, indent=2)
    log.info(f"Sparade {len(stations)} stationer → {path}")


def export_excel(stations: List[Dict], path: str = OUTPUT_EXCEL) -> None:
    try:
        import pandas as pd
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        log.warning("pandas/openpyxl saknas — hoppar över Excel-export. Kör: pip install pandas openpyxl")
        return

    COLUMNS = [
        ("Namn",            "name"),
        ("Stad",            "city"),
        ("Adress",          "address"),
        ("Postnr",          "postcode"),
        ("Lat",             "lat"),
        ("Lng",             "lng"),
        ("Operatör",        "operator"),
        ("HGV-kompatibel",  "hgv_compatible"),
        ("Laddpunkter",     "charging_points"),
        ("Max effekt (kW)", "max_power_kw"),
        ("Kontaktorer",     "connectors"),
        ("Öppettider",      "open_hours"),
        ("Status",          "status"),
        ("Källa",           "source"),
    ]

    rows = []
    for s in stations:
        row = {label: s.get(key, "") for label, key in COLUMNS}
        row["Kontaktorer"]   = ", ".join(s.get("connectors") or [])
        row["HGV-kompatibel"] = "Ja" if s.get("hgv_compatible") else "Nej"
        rows.append(row)

    df = pd.DataFrame(rows, columns=[c for c, _ in COLUMNS])
    df.sort_values(["HGV-kompatibel", "Max effekt (kW)"], ascending=[True, False], inplace=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Laddstationer")
        ws = writer.sheets["Laddstationer"]

        header_fill = PatternFill("solid", fgColor="1A5276")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center")

        hgv_col = next(i for i, (l, _) in enumerate(COLUMNS, 1) if l == "HGV-kompatibel")
        green   = PatternFill("solid", fgColor="C6EFCE")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[hgv_col - 1].value == "Ja":
                for cell in row:
                    cell.fill = green

        COL_WIDTHS = {"Namn": 35, "Adress": 30, "Operatör": 25, "Kontaktorer": 25}
        for i, (label, _) in enumerate(COLUMNS, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = COL_WIDTHS.get(label, 14)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    log.info(f"Excel sparad: {path}")


# ═══════════════════════════════════════════════════════════════════════════════
# HUVUDPROGRAM
# ═══════════════════════════════════════════════════════════════════════════════

def main(ocm_api_key: Optional[str] = None) -> None:
    all_stations: List[Dict] = []

    # ── Källa 1: Overpass per stad ──────────────────────────────────────────
    log.info("=== Källa 1: OpenStreetMap (Overpass) ===")
    for city in CITIES_GEO:
        osm = overpass_query(city)
        all_stations.extend(osm)
        time.sleep(35)  # Overpass rate-limit

    # ── Källa 2: Open Charge Map per stad ───────────────────────────────────
    log.info("\n=== Källa 2: Open Charge Map ===")
    for city in CITIES_GEO:
        ocm = ocm_query(city, api_key=ocm_api_key)
        all_stations.extend(ocm)
        time.sleep(1.0)

    # ── Källa 3: NOBIL (en gång för hela Sverige) ────────────────────────────
    log.info("\n=== Källa 3: NOBIL ===")
    nobil = nobil_query_sweden()
    all_stations.extend(nobil)

    log.info(f"\nRådata: {len(all_stations)} poster")

    # Filtrera till korridoren
    all_stations = filter_corridor(all_stations)
    log.info(f"Efter korridor-filter: {len(all_stations)}")

    # Deduplicera
    all_stations = deduplicate(all_stations, distance_m=100)
    log.info(f"Efter deduplicering: {len(all_stations)} unika stationer")

    # Statistik
    hgv_count = sum(1 for s in all_stations if s.get("hgv_compatible"))
    log.info(f"  → HGV-kompatibla: {hgv_count}")
    log.info(f"  → Generella (EV):  {len(all_stations) - hgv_count}")

    export_json(all_stations)
    export_excel(all_stations)

    log.info(
        f"\n✓ Klar!\n"
        f"  {OUTPUT_JSON}   — används av appen\n"
        f"  {OUTPUT_EXCEL}  — för manuell granskning\n\n"
        f"Kör sedan merge_datasets.py för att slå ihop med lager-data."
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="RouteRider — Laddstationsfinnaren")
    parser.add_argument(
        "--ocm-key",
        metavar="API_KEY",
        help="Open Charge Map API-nyckel (valfritt, ökar rate-limit)",
    )
    args = parser.parse_args()
    main(ocm_api_key=args.ocm_key)
