#!/usr/bin/env python3
"""
RouteRider — Lagerfinnaren
Hittar lager, fabriker och distributionscenter som INTE syns på Allabolag
eftersom de är arbetsställen (filialer/driftsställen) snarare än HQ.

Exempel: Autoexpertens lager i Enköping har sin egna adress men företaget
är registrerat på en helt annan ort. Denna modul hittar den fysiska lokalen.

Datakällor (i prioritetsordning):
  1. Overpass API (OpenStreetMap) — gratis, ingen nyckel
  2. Google Places API            — betalt, kräver API-nyckel, bäst träffsäkerhet
  3. Hitta.se                    — gratis, hittar flerorts-företag

Kör: python warehouse_finder.py
     python warehouse_finder.py --google-key AIza...   (för Google Places)
"""

import argparse
import json
import logging
import re
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests
from bs4 import BeautifulSoup
from geopy.extra.rate_limiter import RateLimiter
from geopy.geocoders import Nominatim

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-8s %(message)s")

# ═══════════════════════════════════════════════════════════════════════════════
# KONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

# Städer med deras ungefärliga koordinater (center-punkt för sökning)
# Används av Overpass och Google Places
CITIES_GEO: List[Dict] = [
    {"name": "Enköping",    "lat": 59.635, "lng": 17.077, "radius_m": 8_000},
    {"name": "Västerås",    "lat": 59.610, "lng": 16.545, "radius_m": 12_000},
    {"name": "Eskilstuna",  "lat": 59.371, "lng": 16.510, "radius_m": 10_000},
    {"name": "Strängnäs",   "lat": 59.378, "lng": 17.031, "radius_m": 7_000},
    {"name": "Katrineholm", "lat": 58.996, "lng": 16.208, "radius_m": 7_000},
    {"name": "Hallsberg",   "lat": 59.066, "lng": 15.111, "radius_m": 6_000},
    {"name": "Örebro",      "lat": 59.275, "lng": 15.213, "radius_m": 12_000},
    {"name": "Södertälje",  "lat": 59.196, "lng": 17.626, "radius_m": 10_000},
    {"name": "Flen",        "lat": 59.058, "lng": 16.588, "radius_m": 5_000},
    {"name": "Nyköping",    "lat": 58.753, "lng": 17.005, "radius_m": 8_000},
    {"name": "Norrköping",  "lat": 58.594, "lng": 16.188, "radius_m": 12_000},
    {"name": "Linköping",   "lat": 58.411, "lng": 15.621, "radius_m": 12_000},
    {"name": "Jönköping",   "lat": 57.782, "lng": 14.161, "radius_m": 12_000},
    {"name": "Skövde",      "lat": 58.391, "lng": 13.845, "radius_m": 8_000},
    {"name": "Falköping",   "lat": 58.180, "lng": 13.551, "radius_m": 7_000},
    {"name": "Ulricehamn",  "lat": 57.793, "lng": 13.420, "radius_m": 6_000},
    {"name": "Borås",       "lat": 57.721, "lng": 12.940, "radius_m": 10_000},
    {"name": "Alingsås",    "lat": 57.930, "lng": 12.533, "radius_m": 7_000},
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; RouteRider/1.0; +https://github.com/routerider)",
    "Accept-Language": "sv-SE,sv;q=0.9",
}

OUTPUT_FILE = "routerider_lager.json"

# ═══════════════════════════════════════════════════════════════════════════════
# KÄLLA 1: OVERPASS API (OpenStreetMap) — Gratis
# ═══════════════════════════════════════════════════════════════════════════════

# Använder backup-endpoint om primär är nere
OVERPASS_URLS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
    "https://maps.mail.ru/osm/tools/overpass/api/interpreter",
]

# Förenklad query — färre predicate = snabbare = mindre timeout-risk
OSM_FILTERS = """
(
  way["building"~"warehouse|industrial|factory"]{bbox};
  node["building"~"warehouse|factory"]{bbox};
  way["shop"="wholesale"]{bbox};
  way["landuse"="industrial"]["name"]{bbox};
  node["landuse"="industrial"]["name"]{bbox};
);
out center tags;
"""


def overpass_query(city: Dict) -> List[Dict]:
    """
    Hämtar industri/lager-objekt från OpenStreetMap inom en given radie.
    Returnerar lista med: name, lat, lng, tags, city
    """
    r = city["radius_m"]
    lat, lng = city["lat"], city["lng"]

    # Bounding box: lat_min, lng_min, lat_max, lng_max
    # 1 grad lat ≈ 111 km, 1 grad lng ≈ 111 km × cos(lat)
    import math
    dlat = r / 111_000
    dlng = r / (111_000 * math.cos(math.radians(lat)))
    bbox = f"({lat-dlat:.5f},{lng-dlng:.5f},{lat+dlat:.5f},{lng+dlng:.5f})"

    query = "[out:json][timeout:60];\n" + OSM_FILTERS.replace("{bbox}", bbox)

    for attempt in range(4):
        wait = [0, 30, 60, 90][attempt]
        if wait:
            log.info(f"  Väntar {wait}s innan försök {attempt+1} för {city['name']}...")
            time.sleep(wait)
        for url in OVERPASS_URLS:
            try:
                resp = requests.post(url, data={"data": query}, timeout=90, headers={"User-Agent": "RouteRider/1.0"})
                resp.raise_for_status()
                data = resp.json()
                break
            except Exception as e:
                log.warning(f"Overpass försök {attempt+1} ({url}) misslyckades för {city['name']}: {e}")
        else:
            if attempt == 3:
                return []
            continue
        break
    else:
        return []

    results = []
    for el in data.get("elements", []):
        tags = el.get("tags", {})
        name = tags.get("name") or tags.get("operator") or tags.get("brand") or ""

        # Koordinater (center för ways/relations)
        if el["type"] == "node":
            elat, elng = el.get("lat"), el.get("lon")
        else:
            center = el.get("center", {})
            elat, elng = center.get("lat"), center.get("lon")

        if not elat:
            continue

        results.append({
            "source":      "openstreetmap",
            "city":        city["name"],
            "name":        name,
            "lat":         round(elat, 6),
            "lng":         round(elng, 6),
            "adress":      tags.get("addr:street", "") + " " + tags.get("addr:housenumber", ""),
            "postnr":      tags.get("addr:postcode", ""),
            "ort":         tags.get("addr:city", city["name"]),
            "typ":         _classify_osm(tags),
            "hemsida":     tags.get("website", tags.get("contact:website", "")),
            "telefon":     tags.get("phone", tags.get("contact:phone", "")),
            "osm_tags":    tags,
            # Fyll i via Allabolag-kopplingen i nästa steg
            "org_nr":      "",
            "omsattning_kr": 0,
            "bransch":     "",
        })

    log.info(f"  Overpass → {city['name']}: {len(results)} objekt")
    return results


def _classify_osm(tags: Dict) -> str:
    """Klassificerar OSM-objekt till en läsbar typ."""
    building  = tags.get("building", "")
    industrial = tags.get("industrial", "")
    name      = (tags.get("name", "") + " " + tags.get("operator", "")).lower()

    if "distribut" in name or "terminal" in name:
        return "distributionscenter"
    if "logistik" in name or "logistics" in name:
        return "logistikcenter"
    if industrial == "factory" or building == "factory":
        return "fabrik"
    if industrial == "warehouse" or building == "warehouse":
        return "lager"
    if building == "industrial":
        return "industrilokal"
    return "industri"


# ═══════════════════════════════════════════════════════════════════════════════
# KÄLLA 2: GOOGLE PLACES API — Kräver API-nyckel
# ═══════════════════════════════════════════════════════════════════════════════

GOOGLE_PLACES_URL = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"

GOOGLE_SEARCH_TERMS = [
    "lager",
    "distributionscenter",
    "logistikcentrum",
    "fabrik",
    "grosslager",
    "terminal",
]

GOOGLE_PLACE_TYPES = ["storage", "moving_company", "warehouse"]  # Google Place-typer


def google_places_search(city: Dict, api_key: str) -> List[Dict]:
    """
    Söker Google Places efter lager/fabriker i en stad.
    Returnerar normaliserade platser i samma format som Overpass.
    """
    results: List[Dict] = []
    seen_ids: set = set()

    for term in GOOGLE_SEARCH_TERMS:
        params = {
            "location":  f"{city['lat']},{city['lng']}",
            "radius":    city["radius_m"],
            "keyword":   term,
            "language":  "sv",
            "key":       api_key,
        }

        while True:
            try:
                resp = requests.get(GOOGLE_PLACES_URL, params=params, timeout=15)
                data = resp.json()
            except Exception as e:
                log.error(f"Google Places fel: {e}")
                break

            for place in data.get("results", []):
                pid = place.get("place_id", "")
                if pid in seen_ids:
                    continue
                seen_ids.add(pid)

                loc = place.get("geometry", {}).get("location", {})
                results.append({
                    "source":        "google_places",
                    "city":          city["name"],
                    "name":          place.get("name", ""),
                    "lat":           loc.get("lat"),
                    "lng":           loc.get("lng"),
                    "adress":        place.get("vicinity", ""),
                    "postnr":        "",
                    "ort":           city["name"],
                    "typ":           _classify_google(place),
                    "hemsida":       "",
                    "telefon":       "",
                    "google_place_id": pid,
                    "org_nr":        "",
                    "omsattning_kr": 0,
                    "bransch":       "",
                })

            next_token = data.get("next_page_token")
            if not next_token:
                break
            params = {"pagetoken": next_token, "key": api_key}
            time.sleep(2)  # Google kräver kort paus innan next_page_token aktiveras

        time.sleep(0.2)

    log.info(f"  Google Places → {city['name']}: {len(results)} objekt")
    return results


def _classify_google(place: Dict) -> str:
    types = place.get("types", [])
    name  = place.get("name", "").lower()
    if "distribut" in name or "terminal" in name:
        return "distributionscenter"
    if "logistik" in name or "logistics" in name:
        return "logistikcenter"
    if "fabrik" in name or "tillverkning" in name or "factory" in name:
        return "fabrik"
    if "storage" in types or "lager" in name:
        return "lager"
    return "industri"


# ═══════════════════════════════════════════════════════════════════════════════
# KÄLLA 3: HITTA.SE — Flerorts-företag
# ═══════════════════════════════════════════════════════════════════════════════

HITTA_SEARCH_URL = "https://www.hitta.se/s%C3%B6k"

HITTA_TERMS = ["lager", "distributionscenter", "grosslager", "logistik", "fabrik"]


def hitta_search(city: Dict) -> List[Dict]:
    """
    Söker Hitta.se per stad + sökterm och extraherar företagslistningar.
    Hitta.se visar filialer/arbetsplatser separat från HQ — precis vad vi vill.
    """
    results: List[Dict] = []
    seen: set = set()

    for term in HITTA_TERMS:
        url    = f"{HITTA_SEARCH_URL}?vad={term}&var={city['name']}"
        page   = 1

        while page <= 5:  # Max 5 sidor per term
            paged_url = url + (f"&sida={page}" if page > 1 else "")
            try:
                resp = requests.get(paged_url, headers=HEADERS, timeout=15)
                soup = BeautifulSoup(resp.text, "lxml")
            except Exception as e:
                log.error(f"Hitta.se fel: {e}")
                break

            # Hitta extraherar liststickor — selector kan behöva justeras
            cards = soup.select(
                ".search-result-item, article.company-card, [data-testid='company-result']"
            )
            if not cards:
                break

            for card in cards:
                name_el = card.select_one("h2, h3, .company-name, [data-testid='company-name']")
                name    = name_el.get_text(strip=True) if name_el else ""

                addr_el = card.select_one(".address, .street-address, [data-testid='address']")
                adress  = addr_el.get_text(strip=True) if addr_el else ""

                phone_el = card.select_one(".phone, [data-testid='phone']")
                telefon  = phone_el.get_text(strip=True) if phone_el else ""

                url_el = card.select_one("a[href]")
                detail_url = url_el["href"] if url_el else ""
                if detail_url and not detail_url.startswith("http"):
                    detail_url = "https://www.hitta.se" + detail_url

                dedup_key = name + adress
                if dedup_key in seen or not name:
                    continue
                seen.add(dedup_key)

                results.append({
                    "source":        "hitta.se",
                    "city":          city["name"],
                    "name":          name,
                    "lat":           None,
                    "lng":           None,
                    "adress":        adress,
                    "postnr":        "",
                    "ort":           city["name"],
                    "typ":           "lager/industri",
                    "hemsida":       "",
                    "telefon":       telefon,
                    "hitta_url":     detail_url,
                    "org_nr":        "",
                    "omsattning_kr": 0,
                    "bransch":       "",
                })

            next_btn = soup.find("a", rel="next") or soup.find("a", string=re.compile("Nästa", re.I))
            if not next_btn:
                break
            page += 1
            time.sleep(1.0)

        time.sleep(1.0)

    log.info(f"  Hitta.se → {city['name']}: {len(results)} objekt")
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# GEOCODA HITTA-RESULTAT (saknar koordinater)
# ═══════════════════════════════════════════════════════════════════════════════

def geocode_missing(locations: List[Dict]) -> List[Dict]:
    """Geocodar poster som saknar lat/lng (t.ex. från Hitta.se)."""
    nom = Nominatim(user_agent="routerider/1.0")
    geo = RateLimiter(nom.geocode, min_delay_seconds=1.1)

    for loc in locations:
        if loc.get("lat"):
            continue
        query = ", ".join(p for p in [loc.get("adress"), loc.get("ort"), "Sverige"] if p)
        try:
            result = geo(query)
            if result:
                loc["lat"] = round(result.latitude, 6)
                loc["lng"] = round(result.longitude, 6)
        except Exception:
            pass

    return locations


# ═══════════════════════════════════════════════════════════════════════════════
# KOPPLA TILL ALLABOLAG (hämta omsättning för hittade lokaler)
# ═══════════════════════════════════════════════════════════════════════════════

ALLABOLAG_SEARCH = "https://www.allabolag.se/vad/{name}/var/{city}"


def enrich_from_allabolag(locations: List[Dict]) -> List[Dict]:
    """
    För varje hittad lokal — försök hitta matchande bolag på Allabolag
    för att få omsättning, org.nr och bransch.

    Matchar på företagsnamn (fuzzy) + ort.
    """
    for loc in locations:
        name = loc.get("name", "").strip()
        city = loc.get("city", "").strip()
        if not name or not city:
            continue

        # Rensa namn för sökning (ta bort "AB", "Sverige" etc.)
        search_name = re.sub(r"\b(AB|HB|KB|AB|Sverige|Sweden|Group|Nordic)\b", "", name, flags=re.I).strip()
        url = f"https://www.allabolag.se/vad/{requests.utils.quote(search_name)}/var/{requests.utils.quote(city)}"

        try:
            resp = requests.get(url, headers=HEADERS, timeout=15)
            soup = BeautifulSoup(resp.text, "lxml")
        except Exception:
            continue

        # Ta första sökträffen
        first = soup.find("a", href=re.compile(r"/\d{6}-\d{4}"))
        if not first:
            time.sleep(0.5)
            continue

        href = first["href"]
        full_url = href if href.startswith("http") else "https://www.allabolag.se" + href

        # Hämta org.nr och detaljer från träffen
        orgnr_match = re.search(r"/(\d{6}-\d{4})", href)
        if orgnr_match:
            loc["org_nr"] = orgnr_match.group(1)
        loc["allabolag_url"] = full_url

        # Omsättning ur listningskortet (undviker ett extra sidanrop)
        omsattning_el = first.find_parent().find(string=re.compile(r"\d[\d\s]+(?:MSEK|TSEK|MKR|TKR|kr)", re.I))
        if omsattning_el:
            loc["omsattning_raw"] = omsattning_el.strip()

        time.sleep(0.8)

    return locations


# ═══════════════════════════════════════════════════════════════════════════════
# DEDUPLICERING — slå ihop träffar från olika källor
# ═══════════════════════════════════════════════════════════════════════════════

def deduplicate(locations: List[Dict], distance_m: float = 200) -> List[Dict]:
    """
    Slår ihop poster som pekar på samma fysiska plats (< distance_m ifrån varandra).
    Prioriterar: google_places > openstreetmap > hitta.se
    """
    import math

    SOURCE_PRIO = {"google_places": 0, "openstreetmap": 1, "hitta.se": 2}

    def dist(a: Dict, b: Dict) -> float:
        """Enkel euklidisk approximation i meter."""
        if not (a.get("lat") and b.get("lat")):
            return float("inf")
        dlat = (a["lat"] - b["lat"]) * 111_000
        dlng = (a["lng"] - b["lng"]) * 111_000 * math.cos(math.radians(a["lat"]))
        return math.sqrt(dlat**2 + dlng**2)

    kept: List[Dict] = []
    for loc in sorted(locations, key=lambda x: SOURCE_PRIO.get(x.get("source", ""), 9)):
        duplicate = False
        for existing in kept:
            if dist(loc, existing) < distance_m:
                # Berika existing med data från loc om det saknas
                for key in ("org_nr", "omsattning_kr", "bransch", "hemsida", "telefon"):
                    if not existing.get(key) and loc.get(key):
                        existing[key] = loc[key]
                duplicate = True
                break
        if not duplicate:
            kept.append(loc)

    return kept


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def export_json(locations: List[Dict], path: str = OUTPUT_FILE) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(locations, f, ensure_ascii=False, indent=2)
    log.info(f"Sparade {len(locations)} lokaler → {path}")


def export_excel_supplement(locations: List[Dict], path: str = "routerider_lager.xlsx") -> None:
    """
    Exporterar lagerfynden till ett eget Excel-ark.
    Kan sedan mergas med routerider_foretag.xlsx i Fas 2.
    """
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment

    COLUMNS = [
        ("Namn",           "name"),
        ("Typ",            "typ"),
        ("Stad",           "city"),
        ("Adress",         "adress"),
        ("Postnr",         "postnr"),
        ("Ort",            "ort"),
        ("Lat",            "lat"),
        ("Lng",            "lng"),
        ("Org.nr",         "org_nr"),
        ("Omsättning (kr)","omsattning_kr"),
        ("Bransch",        "bransch"),
        ("Hemsida",        "hemsida"),
        ("Telefon",        "telefon"),
        ("Källa",          "source"),
        ("Allabolag-URL",  "allabolag_url"),
        ("Anteckningar",   "anteckningar"),
    ]

    rows = [{label: loc.get(key, "") for label, key in COLUMNS} for loc in locations]
    df   = pd.DataFrame(rows, columns=[c for c, _ in COLUMNS])
    df.sort_values(["Stad", "Typ"], inplace=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Lager & Fabriker")
        ws = writer.sheets["Lager & Fabriker"]

        fill = PatternFill("solid", fgColor="2D5986")
        font = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws[1]:
            cell.fill      = fill
            cell.font      = font
            cell.alignment = Alignment(horizontal="center")

        for i, (label, _) in enumerate(COLUMNS, 1):
            col = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[col].width = {"Namn": 35, "Adress": 30, "Allabolag-URL": 40}.get(label, 14)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    log.info(f"Excel sparad: {path}")


# ═══════════════════════════════════════════════════════════════════════════════
# HUVUDPROGRAM
# ═══════════════════════════════════════════════════════════════════════════════

def main(google_api_key: Optional[str] = None) -> None:
    all_locations: List[Dict] = []

    for city in CITIES_GEO:
        log.info(f"\n── {city['name']} ─────────────────────────────────────")

        # 1. Overpass (OpenStreetMap) — alltid
        osm = overpass_query(city)
        all_locations.extend(osm)
        time.sleep(35)  # Overpass rate-limit: max ~1 req/30s

        # 2. Google Places — om API-nyckel finns
        if google_api_key:
            goog = google_places_search(city, google_api_key)
            all_locations.extend(goog)
            time.sleep(1)

        # 3. Hitta.se — alltid
        hitta = hitta_search(city)
        all_locations.extend(hitta)
        time.sleep(2)

    log.info(f"\nTotalt rådata: {len(all_locations)} lokaler")

    # Geocoda poster utan koordinater
    all_locations = geocode_missing(all_locations)

    # Slå ihop dubbletter
    all_locations = deduplicate(all_locations, distance_m=150)
    log.info(f"Efter deduplicering: {len(all_locations)} unika lokaler")

    # Koppla till Allabolag för omsättning/org.nr
    log.info("Kopplar till Allabolag för bolagsdata...")
    all_locations = enrich_from_allabolag(all_locations)

    # Exportera
    export_json(all_locations)
    export_excel_supplement(all_locations)

    log.info(
        f"\n✓ Klar!\n"
        f"  {OUTPUT_FILE}           — JSON (används av appen i Fas 3)\n"
        f"  routerider_lager.xlsx   — Excel för outreach\n\n"
        f"Nästa steg: kör merge_datasets.py för att slå ihop med routerider_foretag.xlsx"
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="RouteRider — Lagerfinnaren")
    parser.add_argument(
        "--google-key",
        metavar="API_KEY",
        help="Google Places API-nyckel (valfritt, förbättrar träffsäkerheten)",
    )
    args = parser.parse_args()
    main(google_api_key=args.google_key)
